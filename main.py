import openpyxl.styles
import pandas as pd
from pathlib import Path
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import excel_tools
import config

folder = Path(__file__).parent

data_path = folder / 'Data'
output_path = folder / 'output'

dfList = []

# 提取并清洗与合并销售文件
df_eor_list = []
df = excel_tools.Concat_file(data_path,df_eor_list)

# 清理重复项并生成异常报告
dup_data = excel_tools.Deduplicate_with_audit(df)
df = dup_data.get('clean_df')
dup_row = dup_data.get('dup_row')
dup_row['异常'] = '数据重复'
df_eor_list.append(dup_row)

# 横向合并门店及商品主数据文件
df = excel_tools.Merge_file(df,data_path)

# 查找匹配不上的异常项并生成异常文件
mismatch_row = df[df['门店名称'].isna() | df['商品类别'].isna() | df['销售人员姓名'].isna()]
mismatch_row['异常'] = '门店或商品或人员编号不匹配'
df = df[~(df['门店名称'].isna() | df['商品类别'].isna() | df['销售人员姓名'].isna())]
df_eor_list.append(mismatch_row)

# 生成转换格式失败的异常数据
format_eor = df[df['数量'].isna() | df['单价'].isna() | df['销售金额'].isna()]
df = df[~(df['数量'].isna() | df['单价'].isna() | df['销售金额'].isna())]
df_eor_list.append(format_eor)

# 提取订单号异常数据
df['订单号'] = df['订单号'].astype(str).str.strip()
pattern = r'^SO\d{5}$'
mask = df['订单号'].str.match(pattern, na=False)

error_df = df[~mask]
error_df['异常'] = '订单号格式异常'

df = df[mask]

df_eor_list.append(error_df)
dfList.append(df)



# 生成商品类别汇总
ProductCategoryAggregation = df.groupby('商品类别').agg(

    订单数量 = ('订单号','size'),
    销售数量 = ('数量','sum'),
    总金额 = ('销售金额','sum'),
    平均销售额 = ('销售金额','mean'),
    平均销售量 = ('数量','mean')

)

dfList.append(ProductCategoryAggregation)

#门店汇总
StoreSummary = df.groupby('门店名称').agg(

    订单数量 = ('订单号','size'),
    销售数量 = ('数量','sum'),
    总金额 = ('销售金额','sum'),
    平均销售额 = ('销售金额','mean'),
    平均销售量 = ('数量','mean')

)

dfList.append(StoreSummary)

#区域汇总
regionSummary = df.groupby('区域').agg(

    订单数量 = ('订单号','size'),
    销售数量 = ('数量','sum'),
    总金额 = ('销售金额','sum'),
    平均销售额 = ('销售金额','mean'),
    平均销售量 = ('数量','mean')

)

dfList.append(regionSummary)

#月度销售汇总
df_copy = df.copy()
df_copy['日期'] = df['日期'].dt.to_period('M')

monthSummary = df_copy.groupby('日期').agg(

    订单数量 = ('订单号','size'),
    销售数量 = ('数量','sum'),
    总金额 = ('销售金额','sum'),
    平均销售额 = ('销售金额','mean'),
    平均销售量 = ('数量','mean')

)

dfList.append(monthSummary)

#销售人员汇总
saleManSummary = df.groupby(['销售人员编号','销售人员姓名']).agg(

    订单数量 = ('订单号','size'),
    销售数量 = ('数量','sum'),
    总金额 = ('销售金额','sum'),
    平均销售额 = ('销售金额','mean'),
    平均销售量 = ('数量','mean')

)

dfList.append(saleManSummary)

# 合并异常数据
df_eor = pd.concat(df_eor_list,ignore_index=True)

dfList.append(df_eor)

# 输出文件
with pd.ExcelWriter(output_path / config.INPUT_FILENAME) as writer:

    i = 0
    for data in dfList:
        if not i:
            data.to_excel(writer,sheet_name=config.SHEET_NAME[0],index=False)
        else:
            data.to_excel(writer,sheet_name=config.SHEET_NAME[i],index=True)

        i += 1

    # 用openpyxl将文件输出成固定格式
    wb = writer.book
    for sheet in config.SHEET_NAME:
        ws = wb[sheet]
        # 插入新行
        ws.insert_rows(1)
        merge_range = excel_tools.Row_range(ws, 1)

        ws['A1'] = sheet
        ws.merge_cells(merge_range)

        title_cell = ws["A1"]
        title_cell.font = Font(bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[2]:
            cell.font = Font(bold=True)

        excel_tools.Column_dimensions_width(ws)

        ws.freeze_panes = 'A3'

        ws.auto_filter.ref = excel_tools.Data_range(ws)

        # 找到金额列
        excel_tools.Reset_money_format(ws)

        # 找到时间列
        excel_tools.Reset_time_format(ws)

        if sheet != config.SHEET_NAME[0]:
            excel_tools.Append_total_row(ws, 3, {'订单数量', '销售数量','数量', '金额', '净额'})
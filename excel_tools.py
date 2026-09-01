import openpyxl.styles
import pandas as pd
from openpyxl.utils import get_column_letter
import config

# 数据清理函数
def Data_clean(df,NanRow):
    # 去除空格
    # 获取所有字符串类型的列（包括 object 和 string）
    str_cols = df.select_dtypes(include=['object', 'string']).columns
    for col in str_cols:
        # 非字符串变为 NaN，然后原值回填
        stripped = df[col].str.strip()
        df[col] = stripped.fillna(df[col])
    # 统一表头
    ReColName(df)
    # 转换类型
    if config.COL_MONEY:
        for money in config.COL_MONEY:
            # 检查数字列空值将它们传入NanRow列表中
            empty_rows = df[df[money].isna() | (df[money].astype(str).str.strip() == '')]
            lenlist = len(NanRow)
            if not empty_rows.empty:
                NanRow.append(empty_rows)
            if len(NanRow) != lenlist:
                if df[money].dtype == object:
                    df[money] = df[money].fillna('')
                    df[money] = df[money].replace('', '0')
                df[money] = df[money].fillna(0)
                NanRow[-1]['异常'] = f'{money}为空值，已填0处理'
            df[money] = df[money].astype(str).str.replace(config.CLEAN_PATTERN, '', regex=True)
            df[money] = pd.to_numeric(df[money], errors='coerce')
    if config.COL_NUM:
        for number in config.COL_NUM:
            # 检查数字列空值将它们传入NanRow列表中
            empty_rows = df[df[number].isna()]
            lenlist = len(NanRow)
            if not empty_rows.empty:
                NanRow.append(empty_rows)
            if len(NanRow) != lenlist:
                df = df[number].fillna(0)
                NanRow[-1]['异常'] = f'{number}为空值，已填0处理'
            df[number] = pd.to_numeric(df[number], errors='coerce')
    # 统一时间格式
    if config.COL_TIME:
        for c_time in config.COL_TIME:
            df[c_time] = df[c_time].astype(str)
            df[c_time] = df[c_time].str.replace('年', '/').str.replace('月', '/').str.replace('日', '')
            df[c_time] = pd.to_datetime(df[c_time], errors='coerce', format='mixed')
    # 返回数据框
    return df

# 检查重复项生成df_eor并删除重复项,通过config.AUDIT_ONLY决定是否退化成只删重复项的函数
def Deduplicate_with_audit(df):
    # 函数返回一个字典，在config.AUDIT_ONLY为T时将dup_row设为None
    dup_mask = df.duplicated(keep='first')

    dup_row = None
    if not config.AUDIT_ONLY:
        dup_row = df[dup_mask]
    return {
        'clean_df':df.drop_duplicates(keep='first'),
        'dup_row':dup_row
    }

# 修改表头名字函数
def ReColName(df):
    cols = df.columns.astype(str).tolist()

    for idx, col in enumerate(cols):
        i = 0

        for key, value in config.COL_NAME_KEYWORD.items():
            if col == key:
                cols[idx] = value
                i = 1
                break

        if not i:
            print(f'{col}未找到匹配列，若含义相同请修改config文件\n')

    df.columns = cols

# 文件纵向合并函数
def Concat_file(folder,NanRow):
    df_list = []
    for file in folder.glob(config.CONCAT_KEY):
        df = pd.read_excel(file, header=0)
        df = Data_clean(df,NanRow)
        df_list.append(df)
    return pd.concat(df_list, ignore_index=True)

# 文件横向合并函数
def Merge_file(df,folder):
    for key, value in config.MERGE_KEY_ON2.items():
        df0 = pd.read_excel(folder / key, header=0)
        df = pd.merge(df, df0,
                      left_on=value.get('left_on'),right_on=value.get('right_on'),how=value.get('how'))
    return df

# 调整列宽函数
def Column_dimensions_width(ws):
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        '对每一列取出其所有的行'
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            '对每一单元格取出他字符数量'
            for cell in row:
                text = str(cell.value)
                lines = text.split('\n')
                current_len = max(len(line) for line in lines)
                '如果长度大于max_len则赋值给max_len'
                if current_len > max_len: max_len = current_len
        '计算列宽'
        adjusted_width = (max_len + 2) * 1.25
        ws.column_dimensions[col_letter].width = adjusted_width


# 获取数据范围函数,返回一个字符串
def Data_range(ws):
    max_row = ws.max_row
    max_col = ws.max_column
    end_column_letter = get_column_letter(max_col)
    data_range = f"A1:{end_column_letter}{max_row}"
    return data_range

# 直接根据工作表对象和行号，返回该行的范围字符串
def Row_range(ws, row_num):
    max_col = ws.max_column
    end_col_letter = get_column_letter(max_col)
    return f"A{row_num}:{end_col_letter}{row_num}"

# 添加合计行
def Append_total_row(ws,start_row,keywords):
    end_row = ws.max_row
    total_row = end_row + 1
    ws.cell(row=total_row, column=1, value="合计")
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        cell_val = ws.cell(row=2, column=col_idx).value
        if cell_val is not None and any(keyword in str(cell_val) for keyword in keywords):
            formula = f"=SUM({col_letter}{start_row}:{col_letter}{end_row})"
            ws.cell(row=total_row, column=col_idx, value=formula)

# 修改金额列格式
def Reset_money_format(ws,column=None):
    target_col = None
    if not column:

        for cell in ws[2]:
            if cell.value and any(keyword in str(cell.value) for keyword in config.MONEY_KEY_WORDS):
                target_col = cell.column
                for row in range(3, ws.max_row + 1):
                    cell = ws.cell(row=row, column=target_col)
                    cell.number_format = config.MONEY_FORMAT[0]

        return None

    for cell in ws[2]:
        if cell.value == column:
            target_col = cell.column
            for row in range(3, ws.max_row + 1):
                cell = ws.cell(row=row, column=target_col)
                cell.number_format = config.MONEY_FORMAT[0]

    return None

# 修改日期列格式
def Reset_time_format(ws,column=None):
    target_col = None
    if not column:

        for cell in ws[2]:
            if cell.value and any(keyword in str(cell.value) for keyword in config.TIME_FORMAT):
                target_col = cell.column
                for row in range(3, ws.max_row + 1):
                    cell = ws.cell(row=row, column=target_col)
                    cell.number_format = config.TIME_FORMAT[0]

        return None

    for cell in ws[2]:
        if cell.value == column:
            target_col = cell.column
            for row in range(3, ws.max_row + 1):
                cell = ws.cell(row=row, column=target_col)
                cell.number_format = config.TIME_FORMAT[0]

    return None
